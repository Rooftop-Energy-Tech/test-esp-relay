import asyncio
import json
import os
import re
import sys
import threading
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from http import HTTPStatus
from urllib.parse import parse_qs, urlparse

import websockets
import websockets.asyncio.server
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

PORT = int(os.environ.get("PORT", 8080))
OUTPUT_DIR = "./output"

gateway_ws: websockets.asyncio.server.ServerConnection | None = None
client_ws: websockets.asyncio.server.ServerConnection | None = None

executor = ThreadPoolExecutor(max_workers=4)
_file_locks: dict[str, threading.Lock] = {}
_file_locks_mutex = threading.Lock()

CONFIG_HEADERS = ["Timestamp", "Event", "BaudRate", "Interval", "SlaveId", "Name", "DeviceBaudRate", "Enabled", "SlaveIdRegister", "Registers"]
DATA_HEADERS = ["Timestamp", "SlaveId", "Name", "Label", "Value"]


def sanitize_name(name: str) -> str:
    return re.sub(r"[^\w]", "_", name).strip("_")


def get_device_filepath(slave_id: int, name: str) -> str:
    return os.path.join(OUTPUT_DIR, f"device_{slave_id}_{sanitize_name(name)}.xlsx")


def _get_file_lock(filepath: str) -> threading.Lock:
    with _file_locks_mutex:
        if filepath not in _file_locks:
            _file_locks[filepath] = threading.Lock()
        return _file_locks[filepath]


def _ensure_workbook(filepath: str) -> Workbook:
    if os.path.exists(filepath):
        return load_workbook(filepath)
    wb = Workbook()
    ws_config: Worksheet = wb.active  # type: ignore[assignment]
    ws_config.title = "Config"
    ws_config.append(CONFIG_HEADERS)
    ws_data: Worksheet = wb.create_sheet("Data")
    ws_data.append(DATA_HEADERS)
    return wb


def _write_data_row_sync(filepath: str, slave_id: int, name: str, timestamp: str, label: str, value) -> None:
    lock = _get_file_lock(filepath)
    with lock:
        wb = _ensure_workbook(filepath)
        wb["Data"].append([timestamp, slave_id, name, label, value])
        wb.save(filepath)


def _write_config_row_sync(filepath: str, timestamp: str, event: str, baud_rate: int | None, interval: int | None, device: dict) -> None:
    lock = _get_file_lock(filepath)
    with lock:
        wb = _ensure_workbook(filepath)
        wb["Config"].append([
            timestamp,
            event,
            baud_rate,
            interval,
            device.get("slaveId"),
            device.get("name"),
            device.get("baudRate"),
            device.get("enabled"),
            device.get("slaveIdRegister"),
            json.dumps(device.get("registers", [])),
        ])
        wb.save(filepath)


async def log_data_message(payload: dict) -> None:
    loop = asyncio.get_event_loop()
    timestamp = datetime.now().isoformat(timespec="seconds")
    for device in payload.get("devices", []):
        slave_id = device.get("slaveId")
        name = device.get("name", "Unknown")
        filepath = get_device_filepath(slave_id, name)
        for reading in device.get("readings", []):
            loop.run_in_executor(
                executor,
                _write_data_row_sync,
                filepath, slave_id, name, timestamp,
                reading.get("label"), reading.get("value"),
            )


async def log_config_message(payload: dict, event: str = "config") -> None:
    loop = asyncio.get_event_loop()
    timestamp = datetime.now().isoformat(timespec="seconds")
    baud_rate = payload.get("baudRate")
    interval = payload.get("interval")
    for device in payload.get("devices", []):
        slave_id = device.get("slaveId")
        name = device.get("name", "Unknown")
        filepath = get_device_filepath(slave_id, name)
        loop.run_in_executor(
            executor,
            _write_config_row_sync,
            filepath, timestamp, event, baud_rate, interval, device,
        )


def parse_request(path: str) -> tuple[str, str | None]:
    parsed = urlparse(path)
    params = parse_qs(parsed.query)
    role = params.get("role", [None])[0]
    return parsed.path, role


async def process_request(connection, request):
    path, role = parse_request(request.path)

    if path == "/health":
        return connection.respond(
            HTTPStatus.OK,
            json.dumps({
                "status": "ok",
                "gateway_connected": gateway_ws is not None,
                "client_connected": client_ws is not None,
            }),
        )
    if path != "/ws":
        return connection.respond(HTTPStatus.NOT_FOUND, "Not found\n")
    if role not in ("gateway", "client"):
        return connection.respond(HTTPStatus.BAD_REQUEST, "Missing or invalid ?role= (gateway|client)\n")


async def handle_connection(ws):
    global gateway_ws, client_ws

    _, role = parse_request(ws.request.path)

    if role == "gateway":
        if gateway_ws is not None:
            print("[relay] Rejecting gateway — one is already connected")
            await ws.close(1008, "Another gateway is already connected")
            return
        gateway_ws = ws
        print("[relay] Gateway connected")
        buffer = ""
        try:
            async for message in ws:
                chunk = message if isinstance(message, str) else message.decode()
                print(f"[relay] chunk ({len(chunk)}B): {repr(chunk[:80])}")
                buffer += chunk
                print(f"[relay] buffer ({len(buffer)}B): {repr(buffer[:80])}")
                try:
                    payload = json.loads(buffer)
                    print(f"[gateway -> relay] {buffer}")
                    if client_ws is not None:
                        await client_ws.send(buffer)
                        print(f"[relay -> client] {buffer}")
                    msg_type = payload.get("type")
                    if msg_type == "data":
                        await log_data_message(payload)
                    elif msg_type == "config":
                        await log_config_message(payload)
                    buffer = ""
                except json.JSONDecodeError as exc:
                    print(f"[relay] incomplete ({exc}), buffering...")
        except websockets.ConnectionClosed:
            pass
        finally:
            gateway_ws = None
            print("[relay] Gateway disconnected")

    elif role == "client":
        if client_ws is not None:
            print("[relay] Rejecting client — one is already connected")
            await ws.close(1008, "Another client is already connected")
            return
        client_ws = ws
        print("[relay] Client connected")
        try:
            async for message in ws:
                print(f"[client -> relay] {message}")
                if gateway_ws is not None:
                    await gateway_ws.send(message)
                    print(f"[relay -> gateway] {message}")
                else:
                    await ws.send(json.dumps({"error": "No gateway connected"}))
        except websockets.ConnectionClosed:
            pass
        finally:
            client_ws = None
            print("[relay] Client disconnected")


async def stdin_loop():
    loop = asyncio.get_event_loop()

    while True:
        line = await loop.run_in_executor(None, sys.stdin.readline)
        if not line:
            break
        text = line.strip()
        if not text:
            continue
        if gateway_ws is None:
            print("[relay] No gateway connected, message dropped")
            continue
        try:
            await gateway_ws.send(text)
            print(f"[stdin -> gateway] {text}")
        except websockets.ConnectionClosed:
            print("[relay] Gateway disconnected while sending")


async def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    async with websockets.serve(
        handle_connection,
        "0.0.0.0",
        PORT,
        process_request=process_request,
        max_size=200 * 1024 * 1024,
    ):
        print(f"Listening on 0.0.0.0:{PORT}")
        await stdin_loop()


if __name__ == "__main__":
    asyncio.run(main())
